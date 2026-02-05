from flask import Flask, render_template, request, jsonify, session, redirect, url_for, send_file
from datetime import datetime
import os
import csv
import io
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import firebase_admin
from firebase_admin import credentials, firestore, auth
import sys
import json
import base64

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'sua_chave_secreta_aqui_mude_para_producao')

# Admin via Firebase Auth
# - Para restringir quais contas podem entrar, defina:
#   ADMIN_EMAILS="admin1@dominio.com,admin2@dominio.com"
#   e/ou ADMIN_EMAIL_DOMAIN="dominio.com"
ADMIN_EMAILS = {
    email.strip().lower()
    for email in os.environ.get('ADMIN_EMAILS', '').split(',')
    if email.strip()
}
ADMIN_EMAIL_DOMAIN = os.environ.get('ADMIN_EMAIL_DOMAIN', '').strip().lower() or None

# Config web do Firebase (para o login no browser).
# Defina via env vars para o admin login funcionar no front-end:
# FIREBASE_API_KEY, FIREBASE_AUTH_DOMAIN, FIREBASE_PROJECT_ID, FIREBASE_APP_ID
FIREBASE_WEB_CONFIG = {
    # Defaults são os valores do teu firebaseConfig (podem ser sobrescritos por env vars)
    'apiKey': os.environ.get('FIREBASE_API_KEY', 'AIzaSyAEvUvbhv2vXj8qa1G6r9S8HSr2cFUv_bM'),
    'authDomain': os.environ.get('FIREBASE_AUTH_DOMAIN', 'studio-7634777517-713ea.firebaseapp.com'),
    'projectId': os.environ.get('FIREBASE_PROJECT_ID', 'studio-7634777517-713ea'),
    'storageBucket': os.environ.get('FIREBASE_STORAGE_BUCKET', 'studio-7634777517-713ea.firebasestorage.app'),
    'messagingSenderId': os.environ.get('FIREBASE_MESSAGING_SENDER_ID', '142898689875'),
    'appId': os.environ.get('FIREBASE_APP_ID', '1:142898689875:web:726d61b0a2590e7e4c93a6'),
    'measurementId': os.environ.get('FIREBASE_MEASUREMENT_ID', 'G-3JZQJD550E'),
}


def _env_bool(name: str, default: bool = False) -> bool:
    value = os.environ.get(name)
    if value is None:
        return default
    return value.strip().lower() in {'1', 'true', 'yes', 'on'}


# Persistência no deploy (serverless): Firestore como fonte de verdade.
FIRESTORE_PRIMARY = _env_bool('FIRESTORE_PRIMARY', default=bool(os.environ.get('VERCEL')))
FIRESTORE_ONLY = _env_bool('FIRESTORE_ONLY', default=bool(os.environ.get('VERCEL')))
AUTO_REBUILD_FIRESTORE_META = _env_bool('AUTO_REBUILD_FIRESTORE_META', default=False)

try:
    FIRESTORE_REBUILD_MAX_DOCS = int(os.environ.get('FIRESTORE_REBUILD_MAX_DOCS', '0') or '0')
except Exception:
    FIRESTORE_REBUILD_MAX_DOCS = 0

FIRESTORE_FEEDBACK_COLLECTION = os.environ.get('FIRESTORE_FEEDBACK_COLLECTION', 'feedback').strip() or 'feedback'
FIRESTORE_META_COLLECTION = os.environ.get('FIRESTORE_META_COLLECTION', '_meta').strip() or '_meta'
FIRESTORE_DAILY_COLLECTION = os.environ.get('FIRESTORE_DAILY_COLLECTION', '_meta_daily').strip() or '_meta_daily'


def _is_admin_email_allowed(email: Optional[str]) -> bool:
    if not email:
        return False
    email_l = email.strip().lower()

    if ADMIN_EMAILS:
        return email_l in ADMIN_EMAILS

    if ADMIN_EMAIL_DOMAIN:
        return email_l.endswith(f"@{ADMIN_EMAIL_DOMAIN}")

    # Se não houver restrição configurada, aceita qualquer conta autenticada
    return True

# Inicializar Firebase
firebase_db = None
FIRESTORE_INIT_ERROR = None
try:
    if not firebase_admin._apps:
        # Preferir credenciais via env var (útil em deploy). Caso contrário, usar o ficheiro no repo.
        service_account_json = os.environ.get('FIREBASE_SERVICE_ACCOUNT_JSON', '').strip()
        service_account_b64 = os.environ.get('FIREBASE_SERVICE_ACCOUNT_JSON_B64', '').strip()

        if service_account_json:
            cred_payload = json.loads(service_account_json)
            cred = credentials.Certificate(cred_payload)
        elif service_account_b64:
            decoded = base64.b64decode(service_account_b64.encode('utf-8')).decode('utf-8')
            cred_payload = json.loads(decoded)
            cred = credentials.Certificate(cred_payload)
        else:
            # Fallback local (dev). Em produção no Vercel, recomenda-se usar env vars.
            cred_path = os.path.join(BASE_DIR, 'studio-7634777517-713ea-firebase-adminsdk-fbsvc-7669723ac0.json')
            cred = credentials.Certificate(cred_path)
        firebase_admin.initialize_app(cred, {
            'databaseURL': 'https://studio-7634777517-713ea.firebaseio.com'
        })
        firebase_db = firestore.client()
        print("✓ Firebase inicializado com sucesso")
except Exception as e:
    FIRESTORE_INIT_ERROR = str(e)
    print(f"⚠ Aviso: Firebase/Firestore não está disponível: {e}")


def _firestore_enabled() -> bool:
    return bool(firebase_db and firebase_admin._apps)


def _firestore_primary_effective() -> bool:
    """True quando queremos Firestore como primário E ele está realmente disponível."""
    return bool(FIRESTORE_PRIMARY and _firestore_enabled())


def _firestore_client():
    if not _firestore_enabled():
        raise RuntimeError('Firestore não disponível (firebase_db=None)')
    return firebase_db


def _fs_feedback_col():
    return _firestore_client().collection(FIRESTORE_FEEDBACK_COLLECTION)


def _fs_meta_doc(name: str):
    return _firestore_client().collection(FIRESTORE_META_COLLECTION).document(name)


def _fs_daily_doc(date_str: str):
    return _firestore_client().collection(FIRESTORE_DAILY_COLLECTION).document(date_str)


def _firestore_rebuild_meta(max_docs: int = 0) -> dict:
    """Reconstroi os contadores em _meta e _meta_daily a partir da coleção feedback."""
    client = _firestore_client()

    overall = {
        'total': 0,
        'muito_satisfeito': 0,
        'satisfeito': 0,
        'insatisfeito': 0,
        'lastId': None,
    }
    daily = {}
    max_id = 0
    scanned = 0

    for doc in _fs_feedback_col().stream():
        scanned += 1
        if max_docs and scanned > max_docs:
            break

        d = doc.to_dict() or {}
        overall['total'] += 1

        try:
            fid = int(d.get('id') or 0)
        except Exception:
            fid = 0
        if fid > max_id:
            max_id = fid

        grau = d.get('grau_satisfacao')
        if grau in {'muito_satisfeito', 'satisfeito', 'insatisfeito'}:
            overall[grau] += 1

        date_str = d.get('data')
        if date_str:
            if date_str not in daily:
                daily[date_str] = {
                    'date': date_str,
                    'total': 0,
                    'muito_satisfeito': 0,
                    'satisfeito': 0,
                    'insatisfeito': 0,
                    'lastId': None,
                }
            daily[date_str]['total'] += 1
            if grau in {'muito_satisfeito', 'satisfeito', 'insatisfeito'}:
                daily[date_str][grau] += 1
            if fid and (daily[date_str]['lastId'] is None or fid > int(daily[date_str]['lastId'] or 0)):
                daily[date_str]['lastId'] = fid

    overall['lastId'] = max_id if max_id else overall['lastId']

    stats_ref = _fs_meta_doc('feedbackStats')
    counters_ref = _fs_meta_doc('counters')

    # Escrever em batches (limite ~500 ops por commit)
    batch = client.batch()
    ops = 0

    batch.set(stats_ref, {
        **overall,
        'updatedAt': firestore.SERVER_TIMESTAMP,
    }, merge=False)
    ops += 1

    # feedbackNextId (se não houver max_id, manter mínimo 1)
    next_id = max(1, int(max_id or 0) + 1)
    batch.set(counters_ref, {
        'feedbackNextId': next_id,
        'updatedAt': firestore.SERVER_TIMESTAMP,
    }, merge=True)
    ops += 1

    # Daily docs
    for date_str, payload in daily.items():
        batch.set(_fs_daily_doc(date_str), {
            **payload,
            'updatedAt': firestore.SERVER_TIMESTAMP,
        }, merge=False)
        ops += 1
        if ops >= 450:
            batch.commit()
            batch = client.batch()
            ops = 0

    if ops:
        batch.commit()

    return {
        'scanned': scanned,
        'total': overall['total'],
        'lastId': overall['lastId'],
        'dates': len(daily),
        'feedbackNextId': next_id,
    }


def _firestore_get_overall_stats_doc() -> dict:
    """Lê o doc _meta/feedbackStats e (opcionalmente) reconstrói se não existir."""
    stats_ref = _fs_meta_doc('feedbackStats')
    snap = stats_ref.get()
    if not snap.exists and AUTO_REBUILD_FIRESTORE_META:
        try:
            _firestore_rebuild_meta(max_docs=FIRESTORE_REBUILD_MAX_DOCS)
        except Exception as e:
            print(f"⚠ Aviso: auto-rebuild de meta falhou: {e}")
        snap = stats_ref.get()
    return snap.to_dict() or {}


def _firestore_compute_next_id_from_existing() -> int:
    """Retorna (max(id)+1) a partir dos docs existentes no Firestore."""
    try:
        query = _fs_feedback_col().order_by('id', direction=firestore.Query.DESCENDING).limit(1)
        for doc in query.stream():
            data = doc.to_dict() or {}
            last_id = int(data.get('id') or 0)
            return max(1, last_id + 1)
    except Exception:
        pass
    return 1


def _firestore_count_query(query) -> int:
    """Conta docs em query. Usa agregação se disponível; caso contrário, stream() e conta."""
    try:
        if hasattr(query, 'count'):
            agg = query.count()
            res = agg.get()
            # google-cloud-firestore devolve uma lista de AggregationResult
            if res:
                # suportar várias versões
                first = res[0]
                if hasattr(first, 'value'):
                    return int(first.value)
                if isinstance(first, dict) and 'value' in first:
                    return int(first['value'])
    except Exception:
        pass

    total = 0
    for _ in query.stream():
        total += 1
    return total


def _firestore_update_meta_for_feedback(transaction, feedback_id: int, feedback_data: dict):
    grau = feedback_data.get('grau_satisfacao')
    date_str = feedback_data.get('data')

    stats_ref = _fs_meta_doc('feedbackStats')
    daily_ref = _fs_daily_doc(date_str)
    counters_ref = _fs_meta_doc('counters')

    inc_common = {
        'total': firestore.Increment(1),
        'lastId': feedback_id,
        'updatedAt': firestore.SERVER_TIMESTAMP,
    }
    if grau in {'muito_satisfeito', 'satisfeito', 'insatisfeito'}:
        inc_common[grau] = firestore.Increment(1)

    transaction.set(stats_ref, inc_common, merge=True)

    daily_payload = {
        'date': date_str,
        'total': firestore.Increment(1),
        'lastId': feedback_id,
        'updatedAt': firestore.SERVER_TIMESTAMP,
    }
    if grau in {'muito_satisfeito', 'satisfeito', 'insatisfeito'}:
        daily_payload[grau] = firestore.Increment(1)

    transaction.set(daily_ref, daily_payload, merge=True)

    # Garantir que o próximo id não fica "atrás" (útil quando ids vêm do SQLite)
    try:
        snap = counters_ref.get(transaction=transaction)
        current_next = None
        if snap.exists:
            current_next = (snap.to_dict() or {}).get('feedbackNextId')
        desired_next = int(feedback_id) + 1
        if not current_next or int(current_next) < desired_next:
            transaction.set(counters_ref, {'feedbackNextId': desired_next}, merge=True)
    except Exception:
        # Melhor esforço: se isto falhar, o gerador ainda tenta auto-inicializar depois
        pass


def _firestore_create_feedback_primary(feedback_data: dict) -> int:
    """Cria feedback no Firestore e devolve um id sequencial (transação)."""
    client = _firestore_client()
    transaction = client.transaction()
    initial_next_id = _firestore_compute_next_id_from_existing()

    @firestore.transactional
    def _txn(transaction):
        counters_ref = _fs_meta_doc('counters')
        snap = counters_ref.get(transaction=transaction)
        next_id = None
        if snap.exists:
            next_id = (snap.to_dict() or {}).get('feedbackNextId')
        if not next_id:
            next_id = initial_next_id

        feedback_id = int(next_id)

        # Reservar o próximo id
        transaction.set(counters_ref, {'feedbackNextId': feedback_id + 1}, merge=True)

        # Criar doc do feedback
        doc_ref = _fs_feedback_col().document(f'feedback_{feedback_id}')
        payload = dict(feedback_data)
        payload['id'] = feedback_id
        payload['createdAt'] = firestore.SERVER_TIMESTAMP
        transaction.set(doc_ref, payload)

        # Atualizar meta (stats/daily)
        _firestore_update_meta_for_feedback(transaction, feedback_id, payload)
        return feedback_id

    return _txn(transaction)

@app.route('/')
def index():
    """Página principal com os botões de feedback"""
    return render_template('index.html')


@app.route('/api/health', methods=['GET'])
def health_check():
    """Health check simples (Firestore init)."""
    firebase_ok = _firestore_enabled()
    storage_mode = 'firestore-only' if FIRESTORE_ONLY else ('firestore-primary' if FIRESTORE_PRIMARY else 'firestore-secondary')

    return jsonify({
        'ok': firebase_ok,
        'storageMode': storage_mode,
        'time': datetime.now().isoformat(),
        'python': sys.version.split(' ')[0],
        'firebase': {
            'initialized': bool(firebase_admin._apps),
            'firestoreAvailable': bool(firebase_db),
            'ok': firebase_ok,
            'projectId': FIREBASE_WEB_CONFIG.get('projectId') or None,
            'initError': FIRESTORE_INIT_ERROR,
        }
    })


@app.route('/api/public/summary', methods=['GET'])
def public_summary():
    """Resumo público para o ecrã principal (sem auth)."""
    try:
        hoje = datetime.now().strftime('%Y-%m-%d')

        if not _firestore_enabled():
            return jsonify({'error': 'Firestore não disponível', 'details': FIRESTORE_INIT_ERROR}), 503

        stats = _firestore_get_overall_stats_doc()
        daily = (_fs_daily_doc(hoje).get().to_dict() or {})

        hoje_result = {
            'muito_satisfeito': int(daily.get('muito_satisfeito') or 0),
            'satisfeito': int(daily.get('satisfeito') or 0),
            'insatisfeito': int(daily.get('insatisfeito') or 0),
        }

        return jsonify({
            'date': hoje,
            'today': hoje_result,
            'todayTotal': sum(hoje_result.values()),
            'total': int(stats.get('total') or 0),
            'lastId': stats.get('lastId'),
            'firebaseAvailable': True,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/feedback', methods=['POST'])
def registrar_feedback():
    """Registra o feedback do usuário no Firestore (Firebase)."""
    try:
        data = request.get_json()
        grau_satisfacao = data.get('grau_satisfacao')
        
        if grau_satisfacao not in ['muito_satisfeito', 'satisfeito', 'insatisfeito']:
            return jsonify({'error': 'Grau de satisfação inválido'}), 400
        
        now = datetime.now()
        data_str = now.strftime('%Y-%m-%d')
        hora_str = now.strftime('%H:%M:%S')
        timestamp_str = now.isoformat()
        
        # Dias da semana em português
        dias_semana = ['Segunda-feira', 'Terça-feira', 'Quarta-feira', 
                       'Quinta-feira', 'Sexta-feira', 'Sábado', 'Domingo']
        dia_semana = dias_semana[now.weekday()]
        
        # Preparar dados do feedback
        feedback_data = {
            'grau_satisfacao': grau_satisfacao,
            'data': data_str,
            'hora': hora_str,
            'dia_semana': dia_semana,
            'timestamp': timestamp_str
        }

        if not _firestore_enabled():
            return jsonify({'error': 'Firestore não disponível', 'details': FIRESTORE_INIT_ERROR}), 503

        feedback_id = _firestore_create_feedback_primary(feedback_data)
        feedback_data['id'] = feedback_id
        
        return jsonify({
            'success': True,
            'message': 'Obrigado pelo seu feedback!',
            'id': feedback_id,
            'firebaseAvailable': True,
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/admin_rocha', methods=['GET', 'POST'])
def admin_login():
    """Página de login do admin"""
    next_url = request.args.get('next')
    # Login do admin agora é via Firebase Authentication (no browser) + verificação do token no backend.
    if request.method == 'POST':
        return render_template(
            'admin_login.html',
            error='O login agora é feito via Firebase Authentication.',
            firebase_web_config=FIREBASE_WEB_CONFIG,
            logout_requested=bool(request.args.get('logout')),
            admin_email_domain=ADMIN_EMAIL_DOMAIN,
            admin_emails=sorted(list(ADMIN_EMAILS)),
            next_url=next_url,
        )
    
    # Se já estiver logado, redireciona para o dashboard
    if session.get('admin_logged_in'):
        return redirect(url_for('admin_dashboard'))
    
    return render_template(
        'admin_login.html',
        firebase_web_config=FIREBASE_WEB_CONFIG,
        logout_requested=bool(request.args.get('logout')),
        admin_email_domain=ADMIN_EMAIL_DOMAIN,
        admin_emails=sorted(list(ADMIN_EMAILS)),
        next_url=next_url,
    )

@app.route('/admin_rocha/dashboard')
def admin_dashboard():
    """Dashboard administrativo"""
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))
    
    return render_template('admin_dashboard.html', admin_email=session.get('admin_email'), firebase_web_config=FIREBASE_WEB_CONFIG)


@app.route('/admin_rocha/tv')
def admin_tv():
    """Modo TV (kiosk) do dashboard."""
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login', next='/admin_rocha/tv'))

    return render_template(
        'admin_tv.html',
        admin_email=session.get('admin_email'),
    )

@app.route('/admin_rocha/logout')
def admin_logout():
    """Logout do admin"""
    session.pop('admin_logged_in', None)
    session.pop('admin_uid', None)
    session.pop('admin_email', None)
    # Redireciona com flag para o front-end fazer signOut() do Firebase, se estiver autenticado.
    return redirect(url_for('admin_login', logout=1))


@app.route('/api/admin/login/firebase', methods=['POST'])
def admin_login_firebase():
    """Cria sessão admin a partir de um Firebase ID token (verificado no backend)."""
    try:
        body = request.get_json(silent=True) or {}
        id_token = body.get('idToken')

        if not id_token:
            return jsonify({'error': 'idToken ausente'}), 400

        if not firebase_admin._apps:
            return jsonify({'error': 'Firebase não inicializado'}), 503

        decoded = auth.verify_id_token(id_token)
        uid = decoded.get('uid')
        email = decoded.get('email')

        if not email and uid:
            user = auth.get_user(uid)
            email = user.email

        if not _is_admin_email_allowed(email):
            return jsonify({'error': 'Conta não autorizada para admin'}), 403

        session['admin_logged_in'] = True
        session['admin_uid'] = uid
        session['admin_email'] = email

        return jsonify({'success': True, 'email': email})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/me', methods=['GET'])
def admin_me():
    """Retorna info da sessão admin atual."""
    if not session.get('admin_logged_in'):
        return jsonify({'loggedIn': False}), 200
    return jsonify({'loggedIn': True, 'email': session.get('admin_email')}), 200

@app.route('/api/admin/stats')
def get_stats():
    """Retorna estatísticas gerais"""
    if not session.get('admin_logged_in'):
        return jsonify({'error': 'Não autorizado'}), 401
    
    try:
        if not _firestore_enabled():
            return jsonify({'error': 'Firestore não disponível', 'details': FIRESTORE_INIT_ERROR}), 503

        stats = _firestore_get_overall_stats_doc()
        total_geral = int(stats.get('total') or 0)

        resultado = {
            'muito_satisfeito': int(stats.get('muito_satisfeito') or 0),
            'satisfeito': int(stats.get('satisfeito') or 0),
            'insatisfeito': int(stats.get('insatisfeito') or 0),
            'total': total_geral,
        }

        percentagens = {
            'muito_satisfeito': 0,
            'satisfeito': 0,
            'insatisfeito': 0,
        }
        if total_geral > 0:
            for key in ['muito_satisfeito', 'satisfeito', 'insatisfeito']:
                percentagens[key] = round((resultado[key] / total_geral) * 100, 2)
        resultado['percentagens'] = percentagens
        return jsonify(resultado)
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/stats/daily')
def get_daily_stats():
    """Retorna estatísticas por dia"""
    if not session.get('admin_logged_in'):
        return jsonify({'error': 'Não autorizado'}), 401
    
    try:
        data_filtro = request.args.get('data')

        if not _firestore_enabled():
            return jsonify({'error': 'Firestore não disponível', 'details': FIRESTORE_INIT_ERROR}), 503

        date_str = data_filtro or datetime.now().strftime('%Y-%m-%d')
        daily = (_fs_daily_doc(date_str).get().to_dict() or {})
        return jsonify({
            'muito_satisfeito': int(daily.get('muito_satisfeito') or 0),
            'satisfeito': int(daily.get('satisfeito') or 0),
            'insatisfeito': int(daily.get('insatisfeito') or 0),
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/stats/comparison')
def get_comparison_stats():
    """Retorna comparação entre dois períodos"""
    if not session.get('admin_logged_in'):
        return jsonify({'error': 'Não autorizado'}), 401
    
    try:
        data1_inicio = request.args.get('data1_inicio')
        data1_fim = request.args.get('data1_fim')
        data2_inicio = request.args.get('data2_inicio')
        data2_fim = request.args.get('data2_fim')
        
        if not all([data1_inicio, data1_fim, data2_inicio, data2_fim]):
            return jsonify({'error': 'Datas inválidas'}), 400
        
        if not _firestore_enabled():
            return jsonify({'error': 'Firestore não disponível', 'details': FIRESTORE_INIT_ERROR}), 503

        if True:

            def _sum_period(start: str, end: str):
                totals = {'muito_satisfeito': 0, 'satisfeito': 0, 'insatisfeito': 0, 'total': 0}
                q = _firestore_client().collection(FIRESTORE_DAILY_COLLECTION).where('date', '>=', start).where('date', '<=', end)
                for doc in q.stream():
                    d = doc.to_dict() or {}
                    for key in ['muito_satisfeito', 'satisfeito', 'insatisfeito']:
                        totals[key] += int(d.get(key) or 0)
                    totals['total'] += int(d.get('total') or 0)
                return totals

            periodo1 = _sum_period(data1_inicio, data1_fim)
            periodo2 = _sum_period(data2_inicio, data2_fim)

            resultado = {
                'periodo1': periodo1,
                'periodo2': periodo2,
            }

            resultado['variacao'] = {}
            for key in ['muito_satisfeito', 'satisfeito', 'insatisfeito']:
                val1 = resultado['periodo1'][key]
                val2 = resultado['periodo2'][key]
                if val1 == 0:
                    variacao = 100 if val2 > 0 else 0
                else:
                    variacao = round(((val2 - val1) / val1) * 100, 2)
                resultado['variacao'][key] = variacao

            return jsonify(resultado)
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/historico')
def get_historico():
    """Retorna o histórico de feedbacks"""
    if not session.get('admin_logged_in'):
        return jsonify({'error': 'Não autorizado'}), 401
    
    try:
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 50))
        offset = (page - 1) * per_page

        # Filtros opcionais (não quebram o comportamento atual)
        q = (request.args.get('q') or '').strip()
        grau = (request.args.get('grau') or '').strip()
        data_inicio = (request.args.get('data_inicio') or '').strip()
        data_fim = (request.args.get('data_fim') or '').strip()
        
        if not _firestore_enabled():
            return jsonify({'error': 'Firestore não disponível', 'details': FIRESTORE_INIT_ERROR}), 503

        if True:

            # Pesquisa por ID exato
            if q.isdigit():
                fid = int(q)
                doc = _fs_feedback_col().document(f'feedback_{fid}').get()
                if not doc.exists:
                    return jsonify({'total': 0, 'page': page, 'per_page': per_page, 'total_pages': 0, 'registros': []})
                data_doc = doc.to_dict() or {}
                registro = {
                    'id': data_doc.get('id'),
                    'grau_satisfacao': data_doc.get('grau_satisfacao'),
                    'data': data_doc.get('data'),
                    'hora': data_doc.get('hora'),
                    'dia_semana': data_doc.get('dia_semana'),
                }
                return jsonify({'total': 1, 'page': page, 'per_page': per_page, 'total_pages': 1, 'registros': [registro]})

            # Para evitar dependência de índices compostos no Firestore, fazemos:
            # - filtros de data no Firestore (campo único)
            # - filtro de grau e ordenação fina em Python
            query = _fs_feedback_col()
            if data_inicio:
                query = query.where('data', '>=', data_inicio)
            if data_fim:
                query = query.where('data', '<=', data_fim)

            # Ordenação principal: por id desc (mais recente primeiro). Se o campo id não existir em docs antigos,
            # o Firestore pode falhar; nesse caso, cai no ordering por data.
            try:
                query = query.order_by('id', direction=firestore.Query.DESCENDING)
                docs_iter = query.stream()
            except Exception:
                query = query.order_by('data', direction=firestore.Query.DESCENDING)
                docs_iter = query.stream()

            all_rows = []
            for doc in docs_iter:
                d = doc.to_dict() or {}
                all_rows.append({
                    'id': d.get('id'),
                    'grau_satisfacao': d.get('grau_satisfacao'),
                    'data': d.get('data'),
                    'hora': d.get('hora'),
                    'dia_semana': d.get('dia_semana'),
                })

            if grau in ['muito_satisfeito', 'satisfeito', 'insatisfeito']:
                all_rows = [r for r in all_rows if r.get('grau_satisfacao') == grau]

            # Ordenação final semelhante ao SQLite (data desc, hora desc) quando possível
            def _sort_key(r):
                return (
                    r.get('data') or '',
                    r.get('hora') or '',
                    int(r.get('id') or 0),
                )

            all_rows.sort(key=_sort_key, reverse=True)

            total = len(all_rows)
            page_rows = all_rows[offset:offset + per_page]

            return jsonify({
                'total': total,
                'page': page,
                'per_page': per_page,
                'total_pages': (total + per_page - 1) // per_page,
                'registros': page_rows,
            })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/export/csv')
@app.route('/api/admin/export/xlsx')
def export_csv():
    """Exporta dados em formato Excel"""
    if not session.get('admin_logged_in'):
        return jsonify({'error': 'Não autorizado'}), 401
    
    try:
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        
        if not _firestore_enabled():
            return jsonify({'error': 'Firestore não disponível', 'details': FIRESTORE_INIT_ERROR}), 503

        if True:

            query = _fs_feedback_col()
            if data_inicio:
                query = query.where('data', '>=', data_inicio)
            if data_fim:
                query = query.where('data', '<=', data_fim)

            registros = []
            for doc in query.stream():
                d = doc.to_dict() or {}
                registros.append({
                    'id': d.get('id'),
                    'grau_satisfacao': d.get('grau_satisfacao'),
                    'data': d.get('data'),
                    'hora': d.get('hora'),
                    'dia_semana': d.get('dia_semana'),
                })

            registros.sort(key=lambda r: (r.get('data') or '', r.get('hora') or '', int(r.get('id') or 0)))
        
        # Criar workbook Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Feedback"
        
        # Cabeçalho com formatação
        header = ['ID', 'Grau de Satisfação', 'Data', 'Hora', 'Dia da Semana']
        ws.append(header)
        
        # Formatar cabeçalho
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Adicionar dados
        grau_map = {
            'muito_satisfeito': 'Muito Satisfeito',
            'satisfeito': 'Satisfeito',
            'insatisfeito': 'Insatisfeito'
        }
        
        for row in registros:
            ws.append([
                row['id'],
                grau_map.get(row['grau_satisfacao'], row['grau_satisfacao']),
                row['data'],
                row['hora'],
                row['dia_semana']
            ])
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 18
        
        # Salvar em memória
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'feedback_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/export/csv-plain')
def export_csv_plain():
    """Exporta dados em CSV (texto). Mantém o /export/csv antigo como Excel."""
    if not session.get('admin_logged_in'):
        return jsonify({'error': 'Não autorizado'}), 401

    try:
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')

        if not _firestore_enabled():
            return jsonify({'error': 'Firestore não disponível', 'details': FIRESTORE_INIT_ERROR}), 503

        if True:
            query = _fs_feedback_col()
            if data_inicio:
                query = query.where('data', '>=', data_inicio)
            if data_fim:
                query = query.where('data', '<=', data_fim)
            registros = []
            for doc in query.stream():
                registros.append(doc.to_dict() or {})
            registros.sort(key=lambda r: (r.get('data') or '', r.get('hora') or '', int(r.get('id') or 0)))

        grau_map = {
            'muito_satisfeito': 'Muito Satisfeito',
            'satisfeito': 'Satisfeito',
            'insatisfeito': 'Insatisfeito'
        }

        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['id', 'grau_satisfacao', 'data', 'hora', 'dia_semana'])
        for row in registros:
            writer.writerow([
                row.get('id') if isinstance(row, dict) else row['id'],
                grau_map.get((row.get('grau_satisfacao') if isinstance(row, dict) else row['grau_satisfacao']),
                             (row.get('grau_satisfacao') if isinstance(row, dict) else row['grau_satisfacao'])),
                (row.get('data') if isinstance(row, dict) else row['data']),
                (row.get('hora') if isinstance(row, dict) else row['hora']),
                (row.get('dia_semana') if isinstance(row, dict) else row['dia_semana']),
            ])

        csv_bytes = output.getvalue().encode('utf-8')
        return send_file(
            io.BytesIO(csv_bytes),
            mimetype='text/csv; charset=utf-8',
            as_attachment=True,
            download_name=f'feedback_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/system')
def admin_system():
    """Info de sistema para o dashboard (requer admin)."""
    if not session.get('admin_logged_in'):
        return jsonify({'error': 'Não autorizado'}), 401

    try:
        if not _firestore_enabled():
            return jsonify({'error': 'Firestore não disponível', 'details': FIRESTORE_INIT_ERROR}), 503

        stats = _firestore_get_overall_stats_doc()
        total = int(stats.get('total') or 0)
        last_id = stats.get('lastId')

        db_size = None

        return jsonify({
            'time': datetime.now().isoformat(),
            'python': sys.version.split(' ')[0],
            'total': total,
            'lastId': last_id,
            'db': {
                'path': None,
                'sizeBytes': db_size,
            },
            'firebase': {
                'initialized': bool(firebase_admin._apps),
                'firestoreAvailable': bool(firebase_db),
                'projectId': FIREBASE_WEB_CONFIG.get('projectId') or None,
            },
            'admin': {
                'email': session.get('admin_email'),
            }
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/firestore/rebuild-meta', methods=['POST'])
def admin_firestore_rebuild_meta():
    """Reconstrói os contadores do Firestore a partir da coleção feedback (requer admin)."""
    if not session.get('admin_logged_in'):
        return jsonify({'error': 'Não autorizado'}), 401
    if not _firestore_enabled():
        return jsonify({'error': 'Firestore não disponível'}), 503

    try:
        body = request.get_json(silent=True) or {}
        max_docs = body.get('maxDocs')
        if max_docs is None:
            max_docs = FIRESTORE_REBUILD_MAX_DOCS
        try:
            max_docs = int(max_docs or 0)
        except Exception:
            max_docs = 0

        result = _firestore_rebuild_meta(max_docs=max_docs)
        return jsonify({'success': True, **result})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/export/txt')
def export_txt():
    """Exporta dados em formato TXT"""
    if not session.get('admin_logged_in'):
        return jsonify({'error': 'Não autorizado'}), 401
    
    try:
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        
        if not _firestore_enabled():
            return jsonify({'error': 'Firestore não disponível', 'details': FIRESTORE_INIT_ERROR}), 503

        if True:
            query = _fs_feedback_col()
            if data_inicio:
                query = query.where('data', '>=', data_inicio)
            if data_fim:
                query = query.where('data', '<=', data_fim)
            registros = []
            for doc in query.stream():
                registros.append(doc.to_dict() or {})
            registros.sort(key=lambda r: (r.get('data') or '', r.get('hora') or '', int(r.get('id') or 0)))
        
        # Criar TXT em memória
        output = io.StringIO()
        output.write('=' * 80 + '\n')
        output.write('RELATÓRIO DE FEEDBACK DE SATISFAÇÃO\n')
        output.write('=' * 80 + '\n\n')
        
        grau_map = {
            'muito_satisfeito': 'Muito Satisfeito',
            'satisfeito': 'Satisfeito',
            'insatisfeito': 'Insatisfeito'
        }
        
        for row in registros:
            if isinstance(row, dict):
                rid = row.get('id')
                rg = row.get('grau_satisfacao')
                rd = row.get('data')
                rh = row.get('hora')
                rds = row.get('dia_semana')
            else:
                rid = row['id']
                rg = row['grau_satisfacao']
                rd = row['data']
                rh = row['hora']
                rds = row['dia_semana']

            output.write(f"ID: {rid}\n")
            output.write(f"Grau de Satisfação: {grau_map.get(rg, rg)}\n")
            output.write(f"Data: {rd}\n")
            output.write(f"Hora: {rh}\n")
            output.write(f"Dia da Semana: {rds}\n")
            output.write('-' * 80 + '\n\n')
        
        output.write(f"\nTotal de registros: {len(registros)}\n")
        output.write(f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n")
        
        # Preparar arquivo para download
        output.seek(0)
        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8')),
            mimetype='text/plain',
            as_attachment=True,
            download_name=f'feedback_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.txt'
        )
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/dates')
def get_available_dates():
    """Retorna as datas disponíveis para filtragem"""
    if not session.get('admin_logged_in'):
        return jsonify({'error': 'Não autorizado'}), 401
    
    try:
        if not _firestore_enabled():
            return jsonify({'error': 'Firestore não disponível', 'details': FIRESTORE_INIT_ERROR}), 503

        q = _firestore_client().collection(FIRESTORE_DAILY_COLLECTION).order_by('date', direction=firestore.Query.DESCENDING)
        dates = []
        for doc in q.stream():
            d = doc.to_dict() or {}
            if d.get('date'):
                dates.append(d['date'])
        return jsonify(dates)
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', '8000'))
    app.run(host='127.0.0.1', port=port, debug=True)
